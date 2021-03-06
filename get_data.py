#!/usr/bin/python
# -*- coding: utf-8 -*-
import sys
from openpyxl import load_workbook
import xlsxwriter
import requests
import pandas as pd
from get_json import get_token
from math import ceil
import time
import datetime
sys.path.append("..")

url = "https://www.iwencai.com/stockpick/cache"
headers = {
    'accept-encoding': "gzip, deflate, br",
    'accept-language': "zh-CN,zh;q=0.9",
    'hexin-v': "Ah96v4m6uLJiMbyugxHsKd7KrniqhHOWjdh3ErFsu04VQDVkuVQDdp2oB2TC",
    'user-agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.26 Safari/537.36 Core/1.63.5702.400 QQBrowser/10.2.1893.400",
    'accept': "application/json, text/javascript, */*; q=0.01",
    'x-requested-with': "XMLHttpRequest",
    'connection': "keep-alive",
    'content-encoding': "gzip",
    'content-type': "multipart/form-data; boundary=----WebKitFormBoundary7MA4YWxkTrZu0gW",
    'cache-control': "no-cache",
    'postman-token': "b10e6bdb-b7bf-1dc2-fdf4-5aa2552ec454"
}


def get_data(query):
    now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    base_path = "E:/iwen/data/" + now   # 保存路径
    # save_path = base_path + ".csv"
    xlsx_path = base_path + ".xlsx"

    writer = pd.ExcelWriter(xlsx_path, engine='xlsxwriter')

    token, total_row = get_token.get_token(query) # 获取本次查询token值
    pages = ceil(int(total_row)/70)
    payload = "------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"token\"\r\n\r\n"+ token +\
              "\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"p\"\r\n\r\n1\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"perpage\"\r\n\r\n70\r\n------" \
                      "WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"showType\"\r\n\r\n[\"\",\"\",\"onTable\",\"onTable\",\"onTable\"]\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW--"
    response = requests.request("POST", url, data=payload.encode(encoding='UTF-8'), headers=headers,timeout =5)
    result = response.json()  # 返回首页结果

    result_title,result_detail,is_merge = get_title(result['title'])  # 表头
    df0 = pd.DataFrame(result_title).T
    df_title = pd.DataFrame(result_detail).T.append(df0)    # 拼接 1

    body = get_body(result['result'])  # 首页表体
    df_body_first = pd.DataFrame.from_dict(body)
    df_all = df_title.append(df_body_first) # 拼接 2

    deal_excel(is_merge, df_all, writer, result_detail,query) #处理写入

    # 获取其他页表体

    if pages >= 2:
        for p in range(2, pages + 1):
            payload = "------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"token\"\r\n\r\n" + token + \
                      "\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"p\"\r\n\r\n" \
                      + str(p) + "\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"perpage\"\r\n\r\n70\r\n------" \
                     "WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"showType\"\r\n\r\n[\"\",\"\",\"onTable\",\"onTable\",\"onTable\"]\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW--"
            time.sleep(3)
            response = requests.request("POST", url, data=payload.encode(encoding='UTF-8'), headers=headers,timeout =10)
            result = response.json() # 返回其他页结果
            other_body = get_body(result['result'])
            df_body_other = pd.DataFrame.from_dict(other_body)  # 其他页表体
            df_all.append(df_body_other)  # 拼接 3
        # df_all.to_excel(writer, sheet_name='data', header=False, index=False)  # 拼接 完再写入Excel
        deal_excel(is_merge, df_all, writer, result_detail,query)  # 处理写入

    writer.close()
    print('xlsx数据导出成功!\n'+'路径：'+ xlsx_path)
    # csv_to_excel(save_path,xlsx_path,query)
# 直接双击文件执行
# while True:
#     get_data(str(input('请输入筛选条件：')))
#     input('继续请按Enter键')


# 获取表头
def get_title(title_result):
    result_title = []
    result_detail = []
    if '{' in str(title_result):
        # print('true')
        is_merge = 1
        for i in title_result:
            if isinstance(i, str):
                j = i.replace('\r', '').replace('<br>', '')
                # result_new.append(j.replace('<br>', ''))
                result_detail.append('')
                result_title.append(j)

            if isinstance(i, dict):
                (key, value), = i.items()
                key = key.replace('\r', '').replace('<br>', '')
                # result_title.append(key + '\n' + value[0])
                result_detail.append(key)
                for d in value[1:]:
                    result_detail.append('')
                for v in value:
                    result_title.append(v)
    else:
        # print('False')
        is_merge = 0
        for i in title_result:
             j = i.replace('\r', '').replace('<br>', '')
             result_title.append(j)


    return result_title,result_detail,is_merge


# 获取表体
def get_body(body_result):
    # # 处理一个单元格下存在多个表格情况
    # print(result[0][4][0]['UID']) # 定位到字典里的值
    body = []
    for solo in body_result:
        new_list = []
        if isinstance(solo, list):
            for zi in solo:
                if isinstance(zi, list):
                    st = ''
                    for value in zi:
                        if isinstance(value, dict):
                            for id in value:
                                st = str(value[id]) + '。' + st
                        else:
                            new_list.append(value)
                    if st == '':
                        pass
                    else:
                        new_list.append(st)
                else:
                    # print(zi)
                    new_list.append(zi)
        else:
            # print(solo)
            new_list.append(solo)
        body.append(new_list)
    return body


# Excel写入与美化
def deal_excel(is_merge, df_all, writer, result_detail, query):
    # 如果存在合并单元格，则执行！
    if is_merge == 1:

        df_all.to_excel(writer, sheet_name='data', header=False, index=False, freeze_panes=(2, 2))  # 拼接 完再写入Excel
        workbook = writer.book
        worksheet = writer.sheets['data']

        cell_format = workbook.add_format({'bold': True})
        worksheet.set_row(1, None, cell_format)  # 设置标题为粗体

        worksheet.autofilter(1,0,1, df_all.shape[1])
        worksheet.filter_column('A', 'x != 300*')  # 过滤 300开头的代码

        worksheet.set_zoom(90)
        workbook.set_properties({'comments': query})

        new_list = []
        text_list = []
        for i in result_detail:
            if i != '':
                id = result_detail.index(i)
                new_list.append(id)
                text_list.append(i)
        # print(new_list)
        # print(text_list)
        for i in new_list:
            for j in range(0, len(new_list)):
                num = new_list[1] - new_list[0] - 1  # 当结果中只有一项是合并单元格时，已知会有问题
                if (j % 2) == 0:
                    merge_format = workbook.add_format({'align': 'center', 'bold': True})
                    merge_format.set_fg_color('#7CCD7C')
                    worksheet.merge_range(0, new_list[j], 0, new_list[j] + num, text_list[j], merge_format)
                else:
                    merge_format = workbook.add_format({'align': 'center', 'bold': True})
                    merge_format.set_fg_color('#7AC5CD')
                    worksheet.merge_range(0, new_list[j], 0, new_list[j] + num, text_list[j], merge_format)
    else:
        df_all.to_excel(writer, sheet_name='data', header=False, index=False, freeze_panes=(1, 1))  # 拼接 完再写入Excel
        workbook = writer.book
        worksheet = writer.sheets['data']

        cell_format = workbook.add_format({'bold': True})
        worksheet.set_row(0, None, cell_format)  # 设置标题为粗体

        worksheet.autofilter(0, 0, 0, df_all.shape[1])
        worksheet.filter_column('A', 'x != 300*')  # 过滤 300开头的代码

        worksheet.set_zoom(90)
        workbook.set_properties({'comments': query})


# 将导出的csv转为xlsx。此方法弃用！！
def csv_to_excel(csv_path,xlsx_path,query):
    # xlsxwriter创建excel默认sheet1 并添加 备注信息
    workbook = xlsxwriter.Workbook(xlsx_path)
    workbook.set_properties({'comments': query})
    workbook.close()

    # CSV文件转XLSX
    writer = pd.ExcelWriter(xlsx_path, engine='openpyxl')
    # pd.set_option('max_colwidth', 20)  # 宽度
    book = load_workbook(writer.path)
    writer.book = book
    condition = ['筛选条件：', query]
    condf = pd.DataFrame(condition).T
    condf.to_excel(writer, '筛选条件', header=False, index=False)  # 添加sheet1   筛选条件
    csv_to_xlsx = pd.read_csv(csv_path, encoding='GBK')
    csv_to_xlsx.to_excel(writer, sheet_name='data',header=False, index=False, freeze_panes=(1, 1))  # 添加sheet2  data
    # 删除由xlsxwriter生成的默认sheet1
    std = book.worksheets
    book.remove(std[0])
    writer.close()
    print('xlsx转换成功！\n保存路径：', xlsx_path)
    # 删除源CSV文件
    # if os.path.exists(save_path):
    #     os.remove(save_path)
    #     print('已删源CSV')

# 程序入口
get_data("预测涨停板")
get_data("量比排名前20")
get_data("换手率排名前20")

get_data("大单净量>0 筹码集中")

# get_data("近2天公告利好")
# get_data("连续3日 dde大单净量大于0.3")
# get_data("业绩预增")
# get_data("连续5日 换手率>7")